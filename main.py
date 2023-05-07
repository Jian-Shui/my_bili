import time
import requests
import pickle as p
import os,sys,glob
import threading
from urllib import request
import openpyxl
#from openpyxl.worksheet.merge import merge_cells
#import xlsxwriter as xw
import hashlib

####初始化
root_path = os.getcwd()
sec_path=root_path+"\\.jsw"
image_path=root_path+"\\image"
####



#dir_path = "/path/to/your/directory"

def get_file(dir_path):
    #####这个函数由ChatGPT倾情提供#####(有问题找ChatGPT)
    png_files = glob.glob(os.path.join(dir_path, "*.png"))# 使用 glob 模块匹配指定目录下的所有 PNG 文件
    file_names = []# 创建一个空的文件名列表
    for png_file in png_files:# 遍历所有 PNG 文件
        file_name = os.path.basename(png_file)# 获取文件名（包含后缀）
        file_name_without_ext = os.path.splitext(file_name)[0]# 去除文件名的后缀（.png）并添加到文件名列表
        file_names.append(file_name_without_ext)
    return file_names#返回值是列表，我觉得



def shamd5(s):#生成hash(简称懒得做字符串整合/bushi)
    #####这个函数由ChatGPT倾情提供#####(有问题找ChatGPT)
    hash_object = hashlib.md5()# 创建一个 hashlib 对象
    hash_object.update(s.encode('utf-8'))# 更新哈希对象的内容
    hash_hex = hash_object.hexdigest()# 获取哈希值的十六进制表示
    #print("字符串哈希值：", hash_hex)
    return hash_hex# 输出哈希值,应该是str吧

#shamd5("https://bibilili.ccoomm/111")这个返回值很棒，有3个0！(所以保留一下这行注释
    

def download_image(url, emo_sha,times=3):#times就是重试次数(失败就递归,直到小于0)
    #####这个函数不完全由ChatGPT倾情提供#####(有问题找ChatGPT,递归有问题找剪水)
    global image_list
    save_path=image_path+"\\"+emo_sha+".png"
    if times<0:
        return -1
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
        with open(save_path, 'wb') as file:
            for chunk in response.iter_content(8192):
                file.write(chunk)
        print(f"成功下载图片到: {save_path}")
        image_list.append(emo_sha)
        Return=0
    except requests.exceptions.HTTPError as e:
        print(f"下载图片失败: {e}\n正在重试")
        Return=-1#网络错误
        time.sleep(3)
        download_image(url, emo_sha,times-1)
    except Exception as e:
        print(f"下载图片失败: {e}")
        Return=1#未知错误
        time.sleep(3)
        download_image(url, emo_sha,times-1)
    return Return
'''
# 输入图片 URL
url = input("请输入图片 URL：")
# 输入保存路径
save_path = input("请输入保存路径和文件名：")

# 调用下载函数
download_image(url, save_path)
'''

##
speed_set=[0,0.1,0.2,0.3,0.4,0.5,0.8,1,1.2,1.5,1.7,2,2.5,3,3.5,4,4.5,5]#最慢的速度是5秒/次
speed=5#即speed_set[5],0.5秒
##
_dm_=[]
_stdm_=[]
_info_=[]
top=[[-1,"X"],[-1,"X"]]
info=[-1,"X"]
wtop=[-1,"X"]

GetCook=True
try:
    load_file = open('user.rep','rb')
    get_list = p.load(load_file)
    load_file.close()
    #_csrf='2d1053e5fc00f5d6b6a4b00302e9abdb'
    _csrf=get_list[0]
    #_cookie='buvid3=9CEAE2E7-69E0-4EE7-89C4-54525AA61F5D13448infoc; LIVE_BUVID=AUTO1516255809771962; CURRENT_BLACKGAP=0; blackside_state=0; buvid4=46B91860-2AA8-13A5-D7EC-CFDB52B77FE925940-022012419-R0CZqzUUvBZlDemKbG6Cpw==; buvid_fp_plain=undefined; nostalgia_conf=-1; is-2022-channel=1; hit-dyn-v2=1; _uuid=987CA510D-CFA1-4C85-E410D-F7FEC7D997C533848infoc; DedeUserID=621088496; DedeUserID__ckMd5=20a368bdc8006707; b_ut=5; i-wanna-go-back=-1; i-wanna-go-feeds=-1; dy_spec_agreed=1; b_nut=100; fingerprint3=b70cef516c1717fd9f9bb76d126ab090; hit-new-style-dyn=0; CURRENT_FNVAL=4048; rpdid=0zbfVGhzo0|U0gUCIpV|3D0|3w1OYzU5; fingerprint=fd5ba672a1abb1550969756fd13636a4; CURRENT_QUALITY=112; SESSDATA=b04dd3a1,1687006612,5f553*c2; bili_jct=2d1053e5fc00f5d6b6a4b00302e9abdb; sid=512ul63l; bp_video_offset_621088496=741491194365739000; PVID=2; b_lsid=8D5B5964_1852CC0E990; innersign=1; buvid_fp=fd5ba672a1abb1550969756fd13636a4'
    _cookie=get_list[1]
except FileNotFoundError:
    print("文件不存在,请更新")
    GetCook=False
    time.sleep(0.2)
    
except:
    print("文件已缺损,请更新")
    GetCook=False
    time.sleep(0.2)
    

image_list=get_file(image_path)#获取我们存了多少弹幕表情(文件名是哈希值),不用try

###验证用户身份
if GetCook:
    get_url = 'https://api.bilibili.com/x/web-interface/nav'
    cookie = {
        'Cookie': _cookie
    }
    response = requests.get(get_url, cookies=cookie)
    udata=response.json()
    if udata["code"]==0:
        uuid=udata["data"]["mid"]
        uname=udata["data"]["uname"]
    else:
        GetCook=False
        print("Cookie失效 ",udata["message"])
###End#

    
#input("PAUSE")

if not GetCook:
    print("!!!!!请注意:当前\"游客模式\",可能出现数据异常或记录不全!!!!!")
else:
    print("欢迎登录\n用户名",uname,"\nuid",uuid)

###
roomid=input("输入直播间ID")#测试例
#roomid=24853527
###

###创建excel文件###
file_name = time.strftime("%Y%m%d-%H%M%S", time.localtime())+".xlsx"
cmd_line="copy "+sec_path+"\\world_template.xlsx "+root_path+"\\"+file_name
#print(cmd_line)
os.system(cmd_line)#复制文件
no_stop=True
def wrt_xlsx(file,GetCook):
    #time.sleep(5)
    begin=0
    global _stdm_
    
    global no_stop
    start_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    workbook = openpyxl.load_workbook(file)
    ws = workbook['Sheet1']
    ws["B2"]=start_time
    ws["B3"]=start_time
    wrt_top=12
    lock = threading.Lock()
    
    while no_stop:
        #print(_stdm_)
        if not _stdm_:
            time.sleep(5)
            continue
        #print("!@@")
        '''
        if wtop==info:#初次写入文本
            begin=0
        
        else:
            for i in _dm_:
                if [i["check"]["ts"],i["check"]["ct"]] == wtop:
                    break
                else:
                    begin+=1
        '''
        
        #lock.acquire()
        for j in range(len(_stdm_)):
            time_point="A"+str(wrt_top)
            #strftime()
            time_tuple = time.strptime(_stdm_[j]["time"], "%Y-%m-%d %H:%M:%S")
            # 将时间元组格式化为你需要的字符串格式
            ws[time_point]=time.strftime("%m-%d %H:%M:%S", time_tuple)
            #time.strftime("%m-%d %H:%M:%S", time.localtime(_stdm_[j]["time"]))
            sender_point="B"+str(wrt_top)
            ws[sender_point]=_stdm_[j]["nickname"]
            #============================
            honor_point="C"+str(wrt_top)
            if _stdm_[j]["medal"]:
                wrt_dtl=_stdm_[j]["medal"][1]+"\n"+str(_stdm_[j]["medal"][0])
            else:
                wrt_dtl="-"
            ws[honor_point]=wrt_dtl
            #============================
            detail_point="D"+str(wrt_top)
            ws[detail_point]=_stdm_[j]["text"]
            wrt_top+=2
            begin+=1
        #lock.release()
        #print(_stdm_)
        lock.acquire()
        _stdm_.clear()
        _info_.clear()
        lock.release()
        begin=0
        workbook.save(file)
        print("!@@成功保存一次")
            
            


baseurl = 'https://api.live.bilibili.com/xlive/web-room/v1/dM/gethistory?roomid='
whurl = baseurl+roomid

def getdm(url,Iscook=False):#str型url传输网址,bool型Iscookie是否使用cookie
    if Iscook:
        response = requests.get(url, cookies=cookie)
    else:
        response = requests.get(url)
    #print(response.json())
    return response.json()


def formatdm(dm,item):#item->0=admin,1=room
    global top
    global info
    num=0#计数，调速用
    if top[item]==info:#证明是第一次抓取，全写入
        for piece in dm:
            if piece["dm_type"]==0:#如果是0，就不是表情
                append_line={
                    "isemo": False,
                    "text": piece["text"],
                    "uid": piece["uid"],
                    "nickname": piece["nickname"],
                    "medal": piece["medal"],#列表
                    "time": piece["timeline"],
                    "rnd": piece["rnd"],#弹幕随机编号吗？
                    "id_str": piece["id_str"],
                    "isadmin": piece["isadmin"],
                    "check": piece["check_info"]
                    }
            else:
                append_line={
                    "isemo": True,
                    "text": piece["text"],
                    "uid": piece["uid"],
                    "nickname": piece["nickname"],
                    "emo": piece["emoticon"],
                    "medal": piece["medal"],#列表
                    "time": piece["timeline"],
                    "rnd": piece["rnd"],#弹幕随机编号吗？
                    "id_str": piece["id_str"],
                    "isadmin": piece["isadmin"],
                    "check": piece["check_info"]
                    }
                emo_sha=shamd5(piece["emoticon"]["url"])
                if not (emo_sha in image_list):#没见过的图片，写了!
                    wrt_path=image_path+"\\"+emo_sha+".png"
                    t1 = threading.Thread(target=download_image,args=(piece["emoticon"]["url"],emo_sha))#多线程执行
                    t1.start()#我们选择沉默(懒得写异步返回值获取了，太难了)
            
            _dm_.append(append_line)#这TM又是字典，提取意义到底是什么(方便后期读取,剪枝再说吧,或者说交给gpt)
            _info_.append(piece)#如果占内存后期再改
            top[item]=[piece["check_info"]["ts"],piece["check_info"]["ct"]]
            
                        


            
    else:#这里记得计数
        num=10
        flg=False
        for piece in dm:
            if flg:
                if piece["dm_type"]==0:#如果是0，就不是表情
                    append_line={
                        "isemo": False,
                        "text": piece["text"],
                        "uid": piece["uid"],
                        "nickname": piece["nickname"],
                        "medal": piece["medal"],#列表
                        "time": piece["timeline"],
                        "rnd": piece["rnd"],#弹幕随机编号吗？
                        "id_str": piece["id_str"],
                        "isadmin": piece["isadmin"],
                        "check": piece["check_info"]
                        }
                else:
                    append_line={
                        "isemo": True,
                        "text": piece["text"],
                        "uid": piece["uid"],
                        "nickname": piece["nickname"],
                        "emo": piece["emoticon"],
                        "medal": piece["medal"],#列表
                        "time": piece["timeline"],
                        "rnd": piece["rnd"],#弹幕随机编号吗？
                        "id_str": piece["id_str"],
                        "isadmin": piece["isadmin"],
                        "check": piece["check_info"]
                        }
                    emo_sha=shamd5(piece["emoticon"]["url"])
                    if not (emo_sha in image_list):#没见过的图片，写了!
                        #wrt_path=image_path+"\\"+emo_sha+".png"
                        t1 = threading.Thread(target=download_image,args=(piece["emoticon"]["url"],emo_sha))#多线程执行
                        t1.start()#我们选择沉默,不做返回值(懒得写异步返回值获取了，太难了)
                
                _dm_.append(append_line)#这TM又是字典，提取意义到底是什么(方便后期读取,剪枝再说吧,或者说交给gpt)
                _info_.append(piece)#如果占内存后期再改
                top[item]=[piece["check_info"]["ts"],piece["check_info"]["ct"]]
                continue
            temp=[piece["check_info"]["ts"],piece["check_info"]["ct"]]
            num-=1#每一个重复弹幕会反向减num，计数
            if temp==top[item]:#找到
                print(num)
                flg=True
                continue
        if flg==False:#跟掉了
            print("Fail to find..")
            top[item]=info#重置top
            formatdm(dm,item)#递归
            return -7#加速
        if num<=0:
            print("Save cpu..")
            return 1#适当减速
        if num>=8:
            print("Be quick..3")
            return -3
        if num>=5:
            print("Be quick..2")
            return -2
        if num>=3:
            print("Be quick..1")
            return -1
            

    return 0#正常速度

wrt_thr = threading.Thread(target=wrt_xlsx,args=(file_name,GetCook))#多线程执行
wrt_thr.start()#该函数是 弹幕写入文件


try:
#if True:#调试缩行用
    while True:
        #print("START.")
        _dm_.clear()#初始化
        _info_.clear()
        danm=getdm(whurl,GetCook)#字典json,未处理
        #print("Get")
        if danm["code"]==0:#首先确定返回码是个好习惯
            admin_danm=danm["data"]["admin"]#请注意这是房主##此时的danm变成了存储一堆json的列表
            room_danm=danm["data"]["room"]#这是房间
            
        else:
            time.sleep(2)#孩子怕极了api速率被封号()
            continue
        speed_dt=formatdm(admin_danm,0)#速度变化量
        speed_dt=min(speed_dt,formatdm(room_danm,1))
        speed+=speed_dt
        if speed<0:
            speed=0
        if speed>=len(speed_set):
            speed=len(speed_set)-1
        ###调试
        
        for i in _dm_:
            print(i["text"])
        ###

        stdm_temp = sorted(_dm_, key=lambda x: x["time"])#排序
        lock = threading.Lock()
        lock.acquire()
        for x in stdm_temp:
            _stdm_.append(x)
        lock.release()
        #print("STDM:",_stdm_)
        _dm_.clear()
        #wrt_xlsx(file_name,GetCook)
        time.sleep(speed_set[speed])

except:
    lock = threading.Lock()
    lock.acquire()
    no_stop=False
    lock.release()
    wrt_thr.join()




